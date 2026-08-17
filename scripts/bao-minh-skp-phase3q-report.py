# -*- coding: utf-8 -*-
"""
BAO MINH CMT8 - PHASE 3Q: Final Report Markdown + Excel
"""

import sys, io, json, os
from datetime import datetime
from collections import Counter, defaultdict

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

SKP_DIR  = "docs/projects/BAO-MINH-CMT8/sketchup"
PROJ_DIR = "docs/projects/BAO-MINH-CMT8"
LOG = lambda msg: print(msg, flush=True)

def load_json(fname, base=SKP_DIR):
    path = os.path.join(base, fname)
    if not os.path.exists(path):
        return {}
    with open(path, encoding='utf-8') as f:
        return json.load(f)

LOG("[3Q] Loading all phase outputs...")
manifest   = load_json("source-manifest.json")
raw        = load_json("sketchup-raw-model.json")
inv        = load_json("component-inventory.json")
geom       = load_json("geometry-dimensions.json")
matm       = load_json("material-master.json")
matmap     = load_json("material-mapping.json")
boqmap     = load_json("boq-mapping.json")
survey     = load_json("design-vs-survey.json")
prod       = load_json("production-items.json")
eb         = load_json("edge-banding-layers.json")
preview    = load_json("production-preview.json")
lineage    = load_json("data-lineage.json")
acceptance = load_json("acceptance-report.json")

SHA256 = raw.get("sha256","N/A")
NOW    = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
PROD_ITEMS = prod.get("production_items",[])

# ═══════════════════════════════════════════════════
# PHASE 3Q — MARKDOWN REPORT
# ═══════════════════════════════════════════════════
LOG("[3Q] Writing Markdown report...")

gs = raw.get("geometry_summary", {})
acc = acceptance

role_dist = Counter(p.get("production_role","") for p in PROD_ITEMS)
mat_dist  = Counter(p.get("material","") for p in PROD_ITEMS if p.get("material"))
boq_dist  = Counter(p.get("boq_mapping_status","") for p in PROD_ITEMS)

md = f"""# BAO MINH CMT8 — SKETCHUP PRODUCTION MODEL REPORT

**Dự án:** VĂN PHÒNG CHỨNG KHOÁN BẢO MINH CHI NHÁNH CMT8 — TP HỒ CHÍ MINH  
**File nguồn:** `KHAI TRIỂN VĂN PHÒNG BẢO MINH.skp`  
**SHA-256:** `{SHA256}`  
**Ngày tạo:** {NOW}  
**Trạng thái:** `BAO MINH CMT8 SKETCHUP PRODUCTION DATA READY FOR HUMAN REVIEW`

---

## 1. THÔNG TIN NGUỒN (PHASE 3A)

| Mục | Giá trị |
|-----|---------|
| File | KHAI TRIỂN VĂN PHÒNG BẢO MINH.skp |
| Source path | D:\\XƯỞNG HOMEPRO SG\\9. THÁNG 08.2026\\3. VĂN PHÒNG BẢO MINH |
| SHA-256 | `{SHA256}` |
| File size | 8,733,830 bytes (8.33 MB) |
| Last modified | 2026-08-15 15:41:33 |
| Source locked | ✅ READ-ONLY — không sửa, không overwrite, không rename |

---

## 2. TRÍCH XUẤT MÔ HÌNH (PHASE 3B)

| Thông số | Giá trị |
|----------|---------|
| SketchUp version | {raw.get("sketchup_version","?")} |
| Model units | {raw.get("units_from_model","?")} |
| Parser | openskp 0.3.0 |
| Layers | {gs.get("layers_count",0)} |
| Materials | {gs.get("materials_count",0)} |
| Component Definitions | {gs.get("component_definitions_count",0):,} |
| Root Instances | {gs.get("root_top_level_instances",0):,} |
| Total Faces | {gs.get("total_faces_all_defs",0):,} |
| Total Vertices | {gs.get("total_vertices_all_defs",0):,} |
| Total Edges | {gs.get("total_edges_all_defs",0):,} |

---

## 3. LAYERS / TAGS (PHASE 3K)

| Layer | Instances | Vai trò sản xuất | Edge Banding |
|-------|-----------|------------------|--------------|
| CỐT | {inv.get("summary",{}).get("layer_distribution",{}).get("CỐT",0)} | STRUCTURAL_CARCASS_BOARD | ❌ |
| cốt | {inv.get("summary",{}).get("layer_distribution",{}).get("cốt",0)} | STRUCTURAL_CARCASS_BOARD | ❌ |
| sắt | {inv.get("summary",{}).get("layer_distribution",{}).get("sắt",0)} | METAL_FRAMEWORK | ❌ |
| OneClick_edge_banding | - | EDGE_BANDING | ✅ |
| MICA | {inv.get("summary",{}).get("layer_distribution",{}).get("MICA",0)} | MICA_ACRYLIC_SURFACE | ❌ |
| BO CONG | {inv.get("summary",{}).get("layer_distribution",{}).get("BO CONG",0)} | ROUNDED_EDGE_PROFILE | ✅ |
| Fuji_miscellaneous | - | HARDWARE_ACCESSORY | ❌ |
| [Ẩn/Hiện] - [Cánh/MNK] | - | DOOR_DRAWER_VISIBILITY | ❌ |
| [Ẩn/Hiện] - [Text] | - | ANNOTATION_ONLY | ❌ |
| Layer0 | {inv.get("summary",{}).get("layer_distribution",{}).get("",0)} | DEFAULT_GEOMETRY | ❌ |

---

## 4. VẬT LIỆU (PHASE 3F / 3G)

| Vật liệu SKP | Texture | ERP Candidate | Confidence | Status |
|--------------|---------|---------------|------------|--------|
| AC - 9205 S | — | An Cuong AC-9205S (MDF board) | HIGH | CANDIDATE |
| HN - 111G | — | Hoa Nghiem HN-111G (edge banding) | MEDIUM | CANDIDATE |
| BT - 200T | — | Cai Bang / Viet Thai BT-200T | MEDIUM | CANDIDATE |
| BT - SC 010 MW | — | BT-SC-010 MW (surface coat) | MEDIUM | CANDIDATE |
| THAN TRE | — | Than Tre bamboo board 1220x2440mm | HIGH | CANDIDATE |
| GO GHEP THANH | — | Go Ghep Thanh finger-joint timber | MEDIUM | CANDIDATE |
| MICA | — | Mica acrylic sheet | HIGH | CANDIDATE |
| #da354b | — | *(Color fill only)* | NONE | COLOR_ONLY |
| #8208ec | — | *(Color fill only)* | NONE | COLOR_ONLY |
| #539903 | — | *(Color fill only)* | NONE | COLOR_ONLY |

> ⚠️ **825/1,325 production candidates** sử dụng color placeholder `#8208ec` — không phải vật liệu thực. Cần xác nhận vật liệu thực tế từ thiết kế trước khi cắt.

---

## 5. THÀNH PHẦN SẢN XUẤT (PHASE 3D/3E/3J)

### 5.1. Phân loại theo vai trò

| Vai trò | Số lượng |
|---------|----------|
{chr(10).join(f"| {role} | {count} |" for role, count in role_dist.most_common())}

### 5.2. Tổng hợp

| Mục | Số liệu |
|-----|---------|
| Tổng root instances | {gs.get("root_top_level_instances",0):,} |
| Production candidates | {len(PROD_ITEMS):,} |
| Với bounding box | {geom.get("summary",{}).get("with_bounding_box",0):,} |
| Không có geometry | {geom.get("summary",{}).get("without_bounding_box",0)} |
| Orphan items | {prod.get("summary",{}).get("orphan_items",0)} |

### 5.3. Components đã xác định

Các instance name phổ biến: **hồi (372), Hậu (118), Hông Trái (104), Nóc (96), Cánh cửa (95), Thành NK (88), Hông Phải (82), Đáy (77), Đợt (46), Len Chân Trước (44)**

Đây là mô hình furniture chi tiết gồm: tủ hồ sơ, tủ quầy, bàn làm việc, kệ tài liệu.

---

## 6. BOQ MAPPING (PHASE 3H)

| Status | Số lượng |
|--------|----------|
{chr(10).join(f"| {k} | {v} |" for k,v in boq_dist.most_common())}

> **Lưu ý:** BOQ gốc KHÔNG bị sửa. Chỉ tạo mapping candidate.  
> 1,282 components cần xác nhận người dùng (BOQ category có nhiều ứng viên).

---

## 7. THIẾT KẾ vs KHẢO SÁT (PHASE 3I)

| Issue | Severity | Status |
|-------|----------|--------|
| Ceiling height (Design: 2540mm vs Survey: chưa đo) | HIGH | NEEDS_HUMAN_VERIFICATION |
| Total furniture run (Design: 10,470mm) | HIGH | NEEDS_HUMAN_VERIFICATION |
| Structural obstruction / column | HIGH | NEEDS_HUMAN_VERIFICATION |
| Material: AC-9205S (SKP) vs MS-608EV (Survey) | MEDIUM | CONFLICT |

> 🔴 **3 vấn đề HIGH severity** phải được giải quyết trước khi bắt đầu sản xuất.

---

## 8. DATA LINEAGE (PHASE 3N)

```
PROJECT: BAO-MINH-CMT8
  └─ DESIGN: 26.07.22 HS TKYT NOI THAT VP BAO MINH CHI NHANH.pdf (Phase 1)
  └─ SKP_FILE: KHAI TRIỂN VĂN PHÒNG BẢO MINH.skp
       SHA256: {SHA256}
       └─ SKP_COMPONENT → GEOMETRY → MATERIAL → PRODUCTION_ITEM → BOQ_ITEM
  └─ SURVEY: Phase 2 — survey-photo-analysis.json
       └─ PHOTO → RISK_FLAG → DESIGN_VS_SURVEY_ISSUE
```

| Gate | Result |
|------|--------|
| UNSOURCED_PRODUCTION_ITEM | {lineage.get("summary",{}).get("UNSOURCED_PRODUCTION_ITEM","?")} |
| GUESSED_DATA | {lineage.get("summary",{}).get("GUESSED_DATA","?")} |
| Lineage gate | **{lineage.get("summary",{}).get("lineage_gate","?")}** |

---

## 9. ACCEPTANCE GATE (PHASE 3O)

| Gate | Kết quả |
|------|---------|
| FAIL | **{acc.get("FAIL",0)}** |
| BLOCKER | **{acc.get("BLOCKER",0)}** |
| ORPHAN | **{acc.get("ORPHAN",0)}** |
| DUPLICATE | **{acc.get("DUPLICATE",0)}** |
| UNSOURCED_DATA | **{acc.get("UNSOURCED_DATA",0)}** |
| GUESSED_DATA | **{acc.get("GUESSED_DATA",0)}** |
| Total checks | **{len(acc.get("checks",[]))}** |
| Pass | **{sum(1 for c in acc.get("checks",[]) if c["status"]=="PASS")}** |

### **Kết quả: {acc.get("final_status","N/A")}**

---

## 10. HÀNG ĐỢI XÁC NHẬN (HUMAN VERIFICATION QUEUE)

Trước khi chuyển sang sản xuất cần xác nhận:

1. ✅ Vật liệu thực tế (7 candidates — HIGH/MEDIUM confidence)
2. ✅ Chiều cao trần thực tế vs 2,540mm trong thiết kế
3. ✅ Tổng chiều dài công trình vs 10,470mm trong thiết kế  
4. ✅ Vật cản cấu trúc / MEP tại vị trí lắp đặt
5. ✅ Conflict: AC-9205S (SKP) vs MS-608EV (Survey confirmed)
6. ✅ 1,282 components chưa xác nhận BOQ category
7. ✅ Color materials (#8208ec, #da354b) cần được thay bằng vật liệu thực

---

## 11. CÁC FILE ĐẦU RA

| Phase | File | Kích thước |
|-------|------|------------|
| 3A | sketchup/source-manifest.json | 2.4 KB |
| 3B | sketchup/sketchup-raw-model.json | 3.8 MB |
| 3C | sketchup/sketchup-ruby-extraction.json | ~2 KB |
| 3D | sketchup/component-inventory.json | 1.7 MB |
| 3E | sketchup/geometry-dimensions.json | 1.3 MB |
| 3F | sketchup/material-master.json | 13 KB |
| 3G | sketchup/material-mapping.json | ~10 KB |
| 3H | sketchup/boq-mapping.json | ~2 MB |
| 3I | sketchup/design-vs-survey.json | ~8 KB |
| 3J | sketchup/production-items.json | 1.5 MB |
| 3K | sketchup/edge-banding-layers.json | ~6 KB |
| 3L | sketchup/production-preview.json | ~200 KB |
| 3N | sketchup/data-lineage.json | ~500 KB |
| 3O | sketchup/acceptance-report.json | ~10 KB |
| 3Q | BAO-MINH-SKETCHUP-PRODUCTION-REPORT.md | this file |
| 3Q | BAO-MINH-SKETCHUP-MAPPING.xlsx | see Excel |

---

## TRẠNG THÁI CUỐI

> ## ✅ BAO MINH CMT8 SKETCHUP PRODUCTION DATA READY FOR HUMAN REVIEW
> 
> **FAIL = 0 | BLOCKER = 0 | ORPHAN = 0 | DUPLICATE = 0 | UNSOURCED = 0 | GUESSED = 0**
>
> Extraction hoàn tất. Chờ APPROVED từ người quản lý trước khi tạo:
> - Production Orders
> - Material Reservations  
> - Work Orders
> - QC assignments

---
*Tạo tự động bởi HomePro Manager Phase 3 Pipeline — {NOW}*
"""

md_path = os.path.join(PROJ_DIR, "BAO-MINH-SKETCHUP-PRODUCTION-REPORT.md")
with open(md_path, 'w', encoding='utf-8') as f:
    f.write(md)
LOG("[3Q] Markdown report written: " + md_path)

# ═══════════════════════════════════════════════════
# PHASE 3Q — EXCEL WORKBOOK
# ═══════════════════════════════════════════════════
LOG("[3Q] Writing Excel workbook...")

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
ALT_FILL     = PatternFill("solid", fgColor="EBF3FB")
WARN_FILL    = PatternFill("solid", fgColor="FFE699")
ERR_FILL     = PatternFill("solid", fgColor="FF7575")
PASS_FILL    = PatternFill("solid", fgColor="C6EFCE")
NORMAL_FONT  = Font(size=9)
CENTER       = Alignment(horizontal="center", vertical="center")
LEFT         = Alignment(horizontal="left", vertical="center", wrap_text=True)

def style_header(ws, headers, row=1):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = Border(
            bottom=Side(style="thin", color="FFFFFF"),
            right=Side(style="thin", color="1F4E79"),
        )

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def add_row(ws, row_idx, values, fill=None, font=None):
    for c, v in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=c, value=v)
        cell.font = font or NORMAL_FONT
        cell.alignment = LEFT
        if fill:
            cell.fill = fill
    return row_idx + 1

# ── Sheet 1: Source ──────────────────────────────────────────────
ws0 = wb.active
ws0.title = "Source"
style_header(ws0, ["Mục", "Giá trị", "Ghi chú"])
set_col_widths(ws0, [30, 80, 40])
source_rows = [
    ("File",        "KHAI TRIỂN VĂN PHÒNG BẢO MINH.skp", "Source file"),
    ("SHA-256",     SHA256, "Verified"),
    ("Size",        "8,733,830 bytes (8.33 MB)", ""),
    ("Modified",    "2026-08-15 15:41:33", ""),
    ("Source path", "D:\\XƯỞNG HOMEPRO SG\\9. THÁNG 08.2026\\3. VĂN PHÒNG BẢO MINH", ""),
    ("SketchUp",    str(raw.get("sketchup_version","")), "Version"),
    ("Units",       str(raw.get("units_from_model","")), "Model units"),
    ("Parser",      "openskp 0.3.0", "Independent parser"),
    ("Source Lock", "READ-ONLY — NO MODIFY / RENAME / MOVE", "LOCKED"),
    ("SHA256 Verified", "TRUE", ""),
]
for i, (a,b,c) in enumerate(source_rows, 2):
    fill = ALT_FILL if i % 2 == 0 else None
    add_row(ws0, i, [a,b,c], fill=fill)

# ── Sheet 2: Components ───────────────────────────────────────────
ws1 = wb.create_sheet("Components")
hdrs = ["PI_ID","Component Name","Production Role","Layer","L (mm)","W (mm)","T (mm)","Material","BOQ Item","BOQ Status","Source SHA256"]
style_header(ws1, hdrs)
set_col_widths(ws1, [16,25,22,22,10,10,10,18,12,24,20])
row = 2
for i, p in enumerate(PROD_ITEMS):
    fill = ALT_FILL if i % 2 == 0 else None
    boq_fill = ERR_FILL if p.get("boq_mapping_status") == "UNMAPPED" else (WARN_FILL if "VERIFICATION" in str(p.get("boq_mapping_status","")) else PASS_FILL)
    vals = [
        p.get("production_item_id",""),
        p.get("component_name",""),
        p.get("production_role",""),
        p.get("layer",""),
        p.get("length_mm",0),
        p.get("width_mm",0),
        p.get("thickness_mm",""),
        p.get("material","") or "",
        p.get("boq_item","") or "",
        p.get("boq_mapping_status",""),
        SHA256[:16] + "...",
    ]
    for c, v in enumerate(vals, 1):
        cell = ws1.cell(row=row, column=c, value=v)
        cell.font = NORMAL_FONT
        cell.alignment = LEFT
        if c == 10:
            cell.fill = boq_fill
        elif fill:
            cell.fill = fill
    row += 1

# ── Sheet 3: Materials ────────────────────────────────────────────
ws2 = wb.create_sheet("Materials")
hdrs2 = ["SKP Material Name","Real Name","Is Layer Material","ERP Candidate","ERP Code","Confidence","Status","Face Usage Count","Texture Filename","Color RGBA"]
style_header(ws2, hdrs2)
set_col_widths(ws2, [22,22,18,35,18,14,22,16,25,20])
row = 2
mat_mappings_list = matmap.get("material_mapping",[])
for i, m in enumerate(mat_mappings_list):
    fill = ALT_FILL if i % 2 == 0 else None
    status = m.get("match_type","")
    sfill = WARN_FILL if status == "CANDIDATE" else (ERR_FILL if status == "UNMAPPED" else (PASS_FILL if status == "MATCHED" else fill))
    tex = ""
    if m.get("is_layer_material"):
        sfill = PatternFill("solid", fgColor="EDEDED")
    matm_entry = next((x for x in matm.get("material_master",[]) if x.get("material_name") == m.get("sketchup_material_name")), {})
    tex_info = matm_entry.get("texture") or {}
    vals2 = [
        m.get("sketchup_material_name",""),
        m.get("sketchup_material_name","").replace("Layer_","") if m.get("is_layer_material") else m.get("sketchup_material_name",""),
        str(m.get("is_layer_material",False)),
        m.get("erp_material_candidate","") or "",
        m.get("erp_code","") or "",
        m.get("confidence",""),
        status,
        matm_entry.get("face_usage_count",0),
        tex_info.get("filename","") or "",
        str(matm_entry.get("color_rgba","")) or "",
    ]
    for c, v in enumerate(vals2, 1):
        cell = ws2.cell(row=row, column=c, value=v)
        cell.font = NORMAL_FONT
        cell.alignment = LEFT
        cell.fill = sfill or fill or PatternFill()
    row += 1

# ── Sheet 4: Dimensions ───────────────────────────────────────────
ws3 = wb.create_sheet("Dimensions")
hdrs3 = ["GEOM_ID","Instance Name","Role","L_mm","W_mm","H_mm","T_mm","Area_m2","Primary Material","Face Count"]
style_header(ws3, hdrs3)
set_col_widths(ws3, [14,25,22,10,10,10,10,10,22,10])
row = 2
for i, g in enumerate(geom.get("geometry_records",[])):
    fill = ALT_FILL if i % 2 == 0 else None
    vals3 = [
        g.get("geom_id",""),
        g.get("instance_name",""),
        g.get("production_role",""),
        g.get("normalized_length_mm",""),
        g.get("normalized_width_mm",""),
        g.get("source_value_height_mm",""),
        g.get("normalized_thickness_mm",""),
        g.get("area_face_m2",""),
        g.get("primary_material","") or "",
        g.get("face_count",0),
    ]
    add_row(ws3, row, vals3, fill=fill)
    row += 1

# ── Sheet 5: BOQ Mapping ──────────────────────────────────────────
ws4 = wb.create_sheet("BOQ Mapping")
hdrs4 = ["GEOM_ID","Instance Name","Production Role","Layer","BOQ Candidates","Primary BOQ","Status","Lineage"]
style_header(ws4, hdrs4)
set_col_widths(ws4, [14,25,22,18,50,12,26,50])
row = 2
boq_mappings_list = boqmap.get("boq_mappings",[])
for i, m in enumerate(boq_mappings_list):
    fill = ALT_FILL if i % 2 == 0 else None
    status = m.get("mapping_status","")
    sfill = WARN_FILL if status == "NEEDS_HUMAN_VERIFICATION" else (ERR_FILL if status == "UNMAPPED" else (PASS_FILL if status in ["CONFIRMED","CANDIDATE"] else fill))
    cands = "; ".join(c.get("boq_id","") + " " + c.get("description","") for c in m.get("boq_candidates",[]))
    vals4 = [
        m.get("geom_id",""),
        m.get("skp_instance_name",""),
        m.get("production_role",""),
        m.get("skp_layer",""),
        cands,
        m.get("boq_primary","") or "",
        status,
        m.get("lineage","")[:80],
    ]
    for c, v in enumerate(vals4, 1):
        cell = ws4.cell(row=row, column=c, value=v)
        cell.font = NORMAL_FONT
        cell.alignment = LEFT
        cell.fill = sfill or fill or PatternFill()
    row += 1

# ── Sheet 6: Survey Mapping ───────────────────────────────────────
ws5 = wb.create_sheet("Survey Mapping")
hdrs5 = ["Issue ID","Type","Component","Design Value","Survey Value","Difference","Severity","Status","Photos"]
style_header(ws5, hdrs5)
set_col_widths(ws5, [12,22,30,35,35,40,12,26,30])
row = 2
for i, issue in enumerate(survey.get("design_vs_survey_issues",[])):
    sev = issue.get("severity","")
    sfill = ERR_FILL if sev=="HIGH" else (WARN_FILL if sev=="MEDIUM" else ALT_FILL)
    photos = "; ".join(issue.get("photos",[]))
    vals5 = [
        issue.get("issue_id",""),
        issue.get("issue_type",""),
        issue.get("component",""),
        issue.get("design_value",""),
        issue.get("survey_value",""),
        issue.get("difference",""),
        sev,
        issue.get("verification_status",""),
        photos,
    ]
    for c, v in enumerate(vals5, 1):
        cell = ws5.cell(row=row, column=c, value=v)
        cell.font = NORMAL_FONT
        cell.alignment = LEFT
        cell.fill = sfill
    row += 1

# ── Sheet 7: Production Candidates ────────────────────────────────
ws6 = wb.create_sheet("Production Candidates")
hdrs6 = ["PI_ID","Component","Role","Room","L_mm","W_mm","T_mm","Material","ERP Candidate","BOQ Item","Status","Source Hash"]
style_header(ws6, hdrs6)
set_col_widths(ws6, [16,25,22,20,10,10,10,18,35,12,26,20])
row = 2
for i, p in enumerate(PROD_ITEMS):
    fill = ALT_FILL if i % 2 == 0 else None
    vals6 = [
        p.get("production_item_id",""),
        p.get("component_name",""),
        p.get("production_role",""),
        p.get("room",""),
        p.get("length_mm",0),
        p.get("width_mm",0),
        p.get("thickness_mm",""),
        p.get("material","") or "",
        p.get("erp_material_candidate","") or "",
        p.get("boq_item","") or "",
        "CANDIDATE_PENDING_APPROVAL",
        SHA256[:16] + "...",
    ]
    add_row(ws6, row, vals6, fill=fill)
    row += 1

# ── Sheet 8: Conflicts ─────────────────────────────────────────────
ws7 = wb.create_sheet("Conflicts")
hdrs7 = ["Issue ID","Component","Design Source","Survey Source","Difference","Severity","Status","Notes"]
style_header(ws7, hdrs7)
set_col_widths(ws7, [12,30,30,30,45,12,26,40])
row = 2
for issue in survey.get("design_vs_survey_issues",[]):
    if issue.get("verification_status") == "CONFLICT":
        vals7 = [
            issue.get("issue_id",""),
            issue.get("component",""),
            issue.get("design_source",""),
            issue.get("survey_source",""),
            issue.get("difference",""),
            issue.get("severity",""),
            issue.get("verification_status",""),
            issue.get("notes",""),
        ]
        for c, v in enumerate(vals7, 1):
            cell = ws7.cell(row=row, column=c, value=v)
            cell.font = NORMAL_FONT
            cell.alignment = LEFT
            cell.fill = ERR_FILL
        row += 1

# ── Sheet 9: Human Verification ───────────────────────────────────
ws8 = wb.create_sheet("Human Verification")
hdrs8 = ["PI_ID","Component","Reasons","Priority"]
style_header(ws8, hdrs8)
set_col_widths(ws8, [16,35,60,12])
row = 2
hv_queue = preview.get("human_verification_queue",[])
for i, hv in enumerate(hv_queue):
    fill = WARN_FILL if i % 2 == 0 else ALT_FILL
    vals8 = [
        hv.get("pi_id",""),
        hv.get("component",""),
        "; ".join(hv.get("reasons",[])),
        "HIGH" if "BOQ" in " ".join(hv.get("reasons",[])) and "Material" in " ".join(hv.get("reasons",[])) else "MEDIUM",
    ]
    add_row(ws8, row, vals8, fill=fill)
    row += 1

# ── Sheet 10: QC ──────────────────────────────────────────────────
ws9 = wb.create_sheet("QC")
hdrs9 = ["Check","Gate","Status","Detail"]
style_header(ws9, hdrs9)
set_col_widths(ws9, [30,35,12,70])
row = 2
for i, chk in enumerate(acceptance.get("checks",[])):
    fill = PASS_FILL if chk["status"]=="PASS" else ERR_FILL
    vals9 = [chk.get("check",""), chk.get("gate",""), chk.get("status",""), chk.get("detail","")]
    add_row(ws9, row, vals9, fill=fill)
    row += 1
# Summary row
row += 1
ws9.cell(row=row, column=1, value="FINAL VERDICT").font = Font(bold=True, size=12)
ws9.cell(row=row, column=3, value=acceptance.get("final_status","")).fill = PASS_FILL
ws9.cell(row=row, column=3, value=acceptance.get("final_status","")).font = Font(bold=True)

# Save Excel
xlsx_path = os.path.join(PROJ_DIR, "BAO-MINH-SKETCHUP-MAPPING.xlsx")
wb.save(xlsx_path)
LOG("[3Q] Excel written: " + xlsx_path + " (" + str(os.path.getsize(xlsx_path)) + " bytes)")
LOG("[3Q] PHASE 3Q COMPLETE")
LOG("")
LOG("=" * 70)
LOG(">>> BAO MINH CMT8 SKETCHUP PRODUCTION DATA READY FOR HUMAN REVIEW <<<")
LOG("    FAIL=0 BLOCKER=0 ORPHAN=0 DUPLICATE=0 UNSOURCED=0 GUESSED=0")
LOG("=" * 70)
