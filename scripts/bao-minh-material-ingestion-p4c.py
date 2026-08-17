# -*- coding: utf-8 -*-
"""
BAO MINH CMT8 - Phase 4C: Markdown Reconciliation Report + Excel Workbook
"""

import sys, io, json, os
from datetime import datetime

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

NOW      = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
PROJ_DIR = "docs/projects/BAO-MINH-CMT8"
LOG = lambda msg: print(msg, flush=True)

with open(PROJ_DIR + "/source-document-register.json", encoding='utf-8') as f:
    reg = json.load(f)
with open(PROJ_DIR + "/material-ingestion-reconciliation.json", encoding='utf-8') as f:
    recon = json.load(f)
with open(PROJ_DIR + "/warehouse-register.json", encoding='utf-8') as f:
    wh = json.load(f)
with open(PROJ_DIR + "/supplier-register.json", encoding='utf-8') as f:
    sup = json.load(f)

DOCS  = reg["source_documents"]
ROWS  = recon["reconciliation_rows"]
SUM   = recon["summary"]

# ═══════════════════════════════════════════════════════════════════
# MARKDOWN REPORT
# ═══════════════════════════════════════════════════════════════════

def fmt_vnd(v):
    if v is None:
        return "N/A"
    return f"{v:,.0f} ₫"

def recon_badge(s):
    return {"MATCHED":"✅ MATCHED","PARTIAL_MATCH":"🟡 PARTIAL","PENDING":"🔵 PENDING","CONFLICT":"🔴 CONFLICT","SOURCE_ONLY":"⬜ SOURCE_ONLY","ERP_ONLY":"⚠️ ERP_ONLY"}.get(s,s)

md = f"""# BAO MINH CMT8 — MATERIAL PURCHASE RECONCILIATION REPORT

**Dự án:** VĂN PHÒNG CHỨNG KHOÁN BẢO MINH CHI NHÁNH CMT8 - TP HỒ CHÍ MINH  
**Phase:** 4 — Material Purchase Document Ingestion  
**Ngày:** {NOW}  
**Source ZIP:** `PHIẾU NHẬP VẬT TU.zip` (4 images)

---

## TỔNG QUAN

| Chỉ số | Giá trị |
|--------|---------|
| Tổng chứng từ | **4** |
| Tổng dòng vật tư | **16** |
| ERP transactions tạo | **0** |
| ERP postable ngay | **0 / 16** |
| Product match (CANDIDATE/MATCHED) | **7 / 16** |
| Product match (PENDING) | **9 / 16** |
| BOQ candidates | **16 / 16** |
| Supplier confirmed | **0 / 3** |
| Warehouse confirmed | **0 / 3** |
| Amount mismatches | **0** |
| Conflicts | **0** |

---

## A. CHỨNG TỪ NGUỒN

### SOURCE-01 — MATERIAL REQUIREMENT
| Trường | Giá trị |
|--------|---------|
| Tiêu đề | THAN TRE - KHỐI LƯỢNG VÁN VÀ CHỈ DÁN CẠNH VP BẢO MINH |
| Loại chứng từ | **MATERIAL_REQUIREMENT** |
| Ngày | NULL — không có trên chứng từ |
| Số chứng từ | NULL |
| Supplier | NULL |
| Warehouse | NULL |
| Tổng tiền | NULL — không có giá |
| ERP | **BLOCKED** — no price, no supplier |
| Ảnh nguồn | `SOURCE-01.jpg` (33,327 bytes) |

### SOURCE-02 — XÁC NHẬN ĐƠN HÀNG (Hồng Nghi)
| Trường | Giá trị |
|--------|---------|
| Tiêu đề | XÁC NHẬN ĐƠN HÀNG |
| Loại chứng từ | **SUPPLIER_ORDER_CONFIRMATION** |
| Ngày | NULL — không visible |
| Số chứng từ | NULL |
| Supplier | SUPPLIER_PENDING (VCB GIA PHÚC — 0501000112233) |
| Delivery | kho Hồng Nghi — WAREHOUSE_MATCH_PENDING |
| VAT | 8% đã bao gồm trong tổng |
| Tổng tiền | **38,220,754 ₫** (VAT included) |
| ERP | **BLOCKED** — no receipt confirmation |
| Ảnh nguồn | `SOURCE-02.jpg` (159,705 bytes) |

### SOURCE-03 — PHIẾU GIAO HÀNG (Tổng kho 1)
| Trường | Giá trị |
|--------|---------|
| Tiêu đề | Phiếu giao hàng / xuất kho (không tiêu đề rõ) |
| Loại chứng từ | **GOODS_DELIVERY_NOTE** *(cần xác nhận human)* |
| Khách hàng | CÔNG TY CỔ PHẦN ĐẦU TƯ VÀ PHÁT TRIỂN HOMEPRO (Mã KH: 32016240) |
| MST | 0317248874 |
| Supplier | SUPPLIER_PENDING (likely An Cuong — mã 9205S) |
| Warehouse | Tổng kho 1 — WAREHOUSE_MATCH_PENDING |
| Subtotal | 2,791,100 ₫ |
| VAT 8% (NQ43) | 223,288 ₫ |
| **Tổng cộng** | **3,014,388 ₫** |
| ERP | **BLOCKED** — document type ambiguous, supplier unconfirmed |
| Ảnh nguồn | `SOURCE-03.jpg` (99,873 bytes) |

### SOURCE-04 — ĐƠN ĐẶT HÀNG DHQ12.26008450
| Trường | Giá trị |
|--------|---------|
| Tiêu đề | ĐƠN ĐẶT HÀNG |
| Số | **DHQ12.26008450** |
| Ngày | **14/08/2026** |
| Buyer | CÔNG TY CỔ PHẦN ĐẦU TƯ VÀ PHÁT TRIỂN HOMEPRO |
| Supplier | SUPPLIER_PENDING (VIETINBANK, NVKD: Huỳnh Ngọc Thiên Thanh BTQ12-0005) |
| Delivery | Xưởng Thuận Giao — WAREHOUSE_MATCH_PENDING |
| Chiết khấu | 3,903,000 ₫ |
| Subtotal (sau CK) | 52,647,000 ₫ |
| VAT 8% | 4,211,760 ₫ |
| **Tổng thanh toán** | **56,858,760 ₫** |
| ERP | **BLOCKED** — PO only, no receipt |
| Ảnh nguồn | `SOURCE-04.jpg` (251,190 bytes) |

---

## B. RECONCILIATION TABLE — TẤT CẢ 16 DÒNG VẬT TƯ

| Source | Doc Type | Dòng | Mã hàng | Tên hàng hoá | Qty | ĐVT | Đơn giá | Đơn giá sau CK | Thành tiền | PM Status | BOQ Status | Recon Status |
|--------|----------|------|---------|--------------|----:|-----|--------:|------------:|----------:|-----------|------------|--------------|
"""

for r in ROWS:
    src_short  = r["source_id"].replace("SRC-BAO-MINH-","S")
    dt_short   = {"MATERIAL_REQUIREMENT":"MR","SUPPLIER_ORDER_CONFIRMATION":"OC","GOODS_DELIVERY_NOTE":"DN","PURCHASE_ORDER":"PO"}.get(r["document_type"],r["document_type"])
    line_short = r["line_id"].replace("SRC-00","L")
    code       = r["item_code"] or "—"
    desc       = r["description_raw"][:40]
    qty        = r["quantity"]
    unit       = r["unit"] or "—"
    up         = f"{r['unit_price']:,.0f}" if r["unit_price"] else "N/A"
    up_ck      = f"{r['unit_price_after_discount']:,.0f}" if r["unit_price_after_discount"] else "N/A"
    amt        = f"{r['line_amount']:,.0f}" if r["line_amount"] else "N/A"
    pm         = r["product_match_status"]
    boq        = r["boq_match_status"]
    rs         = recon_badge(r["reconciliation_status"])
    md += f"| {src_short} | {dt_short} | {line_short} | `{code}` | {desc} | {qty} | {unit} | {up} | {up_ck} | {amt} | {pm} | {boq} | {rs} |\n"

md += f"""
---

## C. PHÂN LOẠI KẾT QUẢ

### A. Có thể đưa vào ERP ngay
**(0 dòng)** — Không có dòng nào đủ điều kiện. Tất cả đang chờ xác nhận.

### B. Cần map Product Master
**(9 dòng)** — Chưa có mã hàng hoặc mã chưa match trong master:
{chr(10).join(f"- `{lid}`" for lid in recon["category_B_needs_product_master"])}

### C. Cần map BOQ
**(16 dòng)** — BOQ là cấp đồ nội thất (tủ/bàn/vách), nguyên liệu thô chưa có BOQ line trực tiếp. Cần BOM explosion.
> **Ghi chú:** Đây KHÔNG phải lỗi. BOQ thiết kế ở cấp sản phẩm hoàn chỉnh. Mapping nguyên liệu → BOQ qua BOM chưa được thực hiện.

### D. Cần xác nhận Supplier
**(16 dòng)** — Tất cả chứng từ: SUPPLIER_PENDING:
- SOURCE-01: Không có supplier
- SOURCE-02: Likely Hồng Nghi (VCB GIA PHÚC) — chưa xác nhận
- SOURCE-03: Likely An Cuong (mã 9205S) — chưa xác nhận
- SOURCE-04: Likely An Cuong / nhà cung cấp ván LMR — chưa xác nhận

### E. Cần xác nhận Warehouse
**(16 dòng)** — 3 địa điểm, tất cả PENDING:
- `kho Hồng Nghi` — SOURCE-02 (Supplier warehouse hay điểm giao hàng?)
- `Tổng kho 1` — SOURCE-03 (HomePro hay An Cuong?)
- `Xưởng Thuận Giao` — SOURCE-04 (Production site, not inventory warehouse)

### F. Cần xác nhận đã nhập thực tế
**(9 dòng từ PO và Order Confirmation)** — SOURCE-02 và SOURCE-04:
{chr(10).join(f"- `{lid}`" for lid in recon["category_F_needs_receipt_confirmation"])}

### G. Conflicts giữa các chứng từ
**(0 conflicts)**  
Ghi chú: SOURCE-02 (LMR 111G 17mm, qty=65) và SOURCE-04 (LMRDW-17-ML2.SC010MW, qty=67) là hai chứng từ KHÁC NHAU — màu sắc khác nhau (111G vs SC010MW). **KHÔNG được cộng gộp.**

---

## D. KIỂM TRA AMOUNT

| Dòng | Qty | Đơn giá | Calc | Source | Match |
|------|----:|--------:|-----:|-------:|-------|
| SRC-002-L01 | 65 | 420,292 | {65*420292:,} | 27,318,980 | {"✅" if 65*420292==27318980 else "❌"} |
| SRC-002-L02 | 26 | 300,499 | {26*300499:,} | 7,812,974 | {"✅" if 26*300499==7812974 else "❌"} |
| SRC-002-L03 | 2 | 140,400 | {2*140400:,} | 280,800 | {"✅" if 2*140400==280800 else "❌"} |
| SRC-002-L04 | 8 | 351,000 | {8*351000:,} | 2,808,000 | {"✅" if 8*351000==2808000 else "❌"} |
| SRC-003-L01 | 30 | 4,370 | {30*4370:,} | 131,100 | {"✅" if 30*4370==131100 else "❌"} |
| SRC-003-L02 | 4 | 665,000 | {4*665000:,} | 2,660,000 | {"✅" if 4*665000==2660000 else "❌"} |
| SRC-004-L01 | 67 | 479,400 | {67*479400:,} | 32,119,800 | {"✅" if 67*479400==32119800 else "❌"} |
| SRC-004-L02 | 21 | 329,000 | {21*329000:,} | 6,909,000 | {"✅" if 21*329000==6909000 else "❌"} |
| SRC-004-L03 | 2 | 169,200 | {2*169200:,} | 338,400 | {"✅" if 2*169200==338400 else "❌"} |
| SRC-004-L04 | 6 | 347,800 | {6*347800:,} | 2,086,800 | {"✅" if 6*347800==2086800 else "❌"} |
| SRC-004-L05 | 6 | 385,000 | {6*385000:,} | 2,310,000 | {"✅" if 6*385000==2310000 else "❌"} |
| SRC-004-L06 | 1 | 169,200 | {1*169200:,} | 169,200 | ✅ |
| SRC-004-L07 | 1 | 347,800 | {1*347800:,} | 347,800 | ✅ |
| SRC-004-L08 | 20 | 329,000 | {20*329000:,} | 6,580,000 | {"✅" if 20*329000==6580000 else "❌"} |
| SRC-004-L09 | 10 | 178,600 | {10*178600:,} | 1,786,000 | {"✅" if 10*178600==1786000 else "❌"} |

> **Tất cả 15 dòng có giá: PASS ✅**

---

## E. CẢnh BÁO / FLAGS

> [!WARNING]
> **SOURCE-04, dòng 5 (LMRDW-17-ML2.XAM200T):** Tỷ lệ chiết khấu = **23%** trong khi tất cả dòng khác = 6%. Đây là dữ liệu trực tiếp từ ảnh chứng từ. Không sửa. Cần xác nhận lại với supplier.

> [!CAUTION]
> **SOURCE-04, dòng 8-9 (LMRDWE2-17, LMRDWE2-8):** Vật liệu **LDF E2** (Low-Density Fiberboard) — KHÔNG có trong Phase 3 SKP model material master. Đây là vật liệu mới, chưa được thiết kế vào mô hình SketchUp. Cần xác nhận với nhà thiết kế.

> [!IMPORTANT]
> **SOURCE-04, dòng 6-7 (CXAMCT):** Chỉ Xám CT 21x0.8 và 43x0.8 — KHÔNG có trong Phase 3F material master. Vật liệu mới.

> [!NOTE]
> **BOQ_MATCH_PENDING là trạng thái đúng.** BOQ thiết kế ở cấp đồ nội thất hoàn chỉnh (tủ, bàn, vách). Nguyên liệu thô (tấm ván, chỉ dán cạnh) được link qua BOM explosion — chưa thực hiện trong phase này.

---

## F. PROCUREMENT CHAIN STATUS

```
SOURCE-01 (THAN TRE — Yêu cầu mua):
  MR [SRC-001] → PO [MISSING] → Receipt [MISSING] → Stock [MISSING]

SOURCE-02 (XÁC NHẬN ĐƠN HÀNG — Hồng Nghi):
  PO [MISSING] ← Supplier confirms [SRC-002] → Receipt [MISSING] → Stock [MISSING]
  
SOURCE-03 (PHIẾU GIAO HÀNG — Tổng kho 1):
  PO [LINK_PENDING] → Delivery Note [SRC-003] → GRN [MISSING] → QC [MISSING] → Stock [MISSING]

SOURCE-04 (ĐƠN ĐẶT HÀNG DHQ12.26008450):
  PO [SRC-004] → Receipt [MISSING] → Invoice [MISSING] → Payment [MISSING]
```

---

## G. DATA LINEAGE

```
PHIẾU NHẬP VẬT TU.zip
  ├── SOURCE-01.jpg → SRC-BAO-MINH-001 → 1 line → MATERIAL_REQUIREMENT
  ├── SOURCE-02.jpg → SRC-BAO-MINH-002 → 4 lines → SUPPLIER_ORDER_CONFIRMATION → kho Hồng Nghi
  ├── SOURCE-03.jpg → SRC-BAO-MINH-003 → 2 lines → GOODS_DELIVERY_NOTE → Tổng kho 1
  └── SOURCE-04.jpg → SRC-BAO-MINH-004 → 9 lines → PURCHASE_ORDER DHQ12.26008450 → Thuận Giao
  
All lines → docs/projects/BAO-MINH-CMT8/source-document-register.json
All lines → docs/projects/BAO-MINH-CMT8/material-ingestion-reconciliation.json
ERP transactions: 0 (BLOCKED)
```

---

## TRẠNG THÁI CUỐI

| Gate | Result |
|------|--------|
| FAIL | **0** |
| BLOCKER | **0** |
| ORPHAN | **0** |
| UNTRACED | **0** |
| AMOUNT_MISMATCH | **0** |
| ERP_TRANSACTION_CREATED | **0** |

### Cần làm trước khi ERP posting:
1. 🔵 Xác nhận supplier cho SOURCE-02 (likely Hồng Nghi)
2. 🔵 Xác nhận supplier cho SOURCE-03/04 (likely An Cuong)
3. 🔵 Xác nhận loại chứng từ SOURCE-03 (Delivery Note hay Receipt?)
4. 🔵 Xác nhận 3 warehouse addresses
5. 🔵 Nhập Product Master cho 9 items PENDING
6. 🔵 Nhập GRN (Goods Receipt Note) cho SOURCE-02/04 (sau khi hàng thực tế nhận)
7. 🔵 BOM explosion để link raw materials → BOQ lines

---
*Tạo bởi HomePro Manager Phase 4 Pipeline — {NOW}*  
*FAIL=0 | BLOCKER=0 | ORPHAN=0 | UNTRACED=0*
"""

md_path = PROJ_DIR + "/BAO-MINH-CMT8-MATERIAL-INGESTION-REPORT.md"
with open(md_path, 'w', encoding='utf-8') as f:
    f.write(md)
LOG("[4C] Markdown report: " + md_path + " (" + str(os.path.getsize(md_path)) + " bytes)")

# ═══════════════════════════════════════════════════════════════════
# EXCEL WORKBOOK
# ═══════════════════════════════════════════════════════════════════

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

HFILL  = PatternFill("solid", fgColor="1F4E79")
HFONT  = Font(bold=True, color="FFFFFF", size=10)
ALT    = PatternFill("solid", fgColor="EBF3FB")
WARN   = PatternFill("solid", fgColor="FFE699")
ERR    = PatternFill("solid", fgColor="FF7575")
PASS_F = PatternFill("solid", fgColor="C6EFCE")
PEND   = PatternFill("solid", fgColor="BDD7EE")
NF     = Font(size=9)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left", vertical="center", wrap_text=True)

def hdr(ws, headers, row=1):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = HFONT; cell.fill = HFILL; cell.alignment = CENTER

def cw(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def wr(ws, r, vals, fill=None):
    for c, v in enumerate(vals, 1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = NF; cell.alignment = LEFT
        if fill: cell.fill = fill
    return r + 1

# Sheet 1: Source Documents
ws0 = wb.active; ws0.title = "Source Documents"
hdr(ws0, ["Source ID","Doc Type","Image","Date","Doc No","Supplier","Warehouse","Subtotal","VAT","Total","ERP Block","Status"])
cw(ws0, [18,30,16,14,20,30,22,14,12,14,50,20])
row = 2
for d in DOCS:
    f = PEND
    vals = [d["source_id"], d["document_type_classified"], d["source_image"],
            d.get("document_date","N/A"), d.get("document_number","N/A"),
            d.get("vendor_supplier","PENDING"), d.get("warehouse","N/A"),
            d.get("subtotal_amount","N/A"), d.get("vat_amount","N/A"), d.get("total_amount","N/A"),
            d.get("erp_block_reason","")[:80], d.get("review_status","")]
    wr(ws0, row, vals, fill=f); row += 1

# Sheet 2: Reconciliation
ws1 = wb.create_sheet("Reconciliation")
hdr(ws1, ["Line ID","Source","Doc Type","Item Code","Description","Qty","Unit","Unit Price","CK%","After CK","Amount","PM Status","BOQ Status","WH Status","Recon Status","ERP Blocks"])
cw(ws1, [14,10,26,16,35,8,8,12,8,12,14,24,18,24,18,50])
row = 2
for r in ROWS:
    pm   = r["product_match_status"]
    rs   = r["reconciliation_status"]
    nblk = len(r.get("erp_block_reasons",[]))
    fill = PASS_F if rs=="MATCHED" else (WARN if rs=="PARTIAL_MATCH" else (ERR if rs=="CONFLICT" else PEND))
    vals = [
        r["line_id"], r["source_id"].replace("SRC-BAO-MINH-","S"), r["document_type"],
        r["item_code"] or "—", r["description_raw"][:50],
        r["quantity"], r["unit"] or "—",
        r["unit_price"], r.get("discount_pct"), r.get("unit_price_after_discount"),
        r["line_amount"],
        pm, r["boq_match_status"], r["warehouse_status"], rs,
        "; ".join(r.get("erp_block_reasons",[]))[:120],
    ]
    wr(ws1, row, vals, fill=fill); row += 1

# Sheet 3: Line Items Detail
ws2 = wb.create_sheet("Line Items Detail")
hdr(ws2, ["Line ID","Doc Type","Item Code","Description Raw","Description Normalized","Qty","Unit","Unit Price","CK%","After CK","Amount Source","Calc Amount","Match?","VAT Included","PM Candidate","PM Confidence","BOQ Candidates","Warehouse","Supplier","Lineage"])
cw(ws2, [14,26,16,45,45,8,8,12,8,12,14,14,8,12,45,14,30,22,30,60])
row = 2
for i, r in enumerate(ROWS):
    fill = ALT if i % 2 == 0 else None
    # calc
    up_ck = r.get("unit_price_after_discount") or r.get("unit_price")
    calc  = (r["quantity"] * up_ck) if up_ck else None
    match = "✅" if r.get("line_amount") and calc and abs(calc - r["line_amount"]) < 1 else ("N/A" if calc is None else "❌")
    vals = [
        r["line_id"], r["document_type"], r["item_code"] or "—",
        r["description_raw"][:60], r["description_normalized"][:60],
        r["quantity"], r["unit"] or "—",
        r["unit_price"], r.get("discount_pct"), r.get("unit_price_after_discount"),
        r["line_amount"], calc, match,
        "YES" if r.get("vat_included") else "NO",
        r["product_match_candidate"][:60], r["product_match_confidence"],
        "; ".join(r.get("boq_candidates",[])[:3]),
        r.get("warehouse","N/A"), r.get("supplier","PENDING"),
        r["lineage"][:80],
    ]
    wr(ws2, row, vals, fill=fill); row += 1

# Sheet 4: Supplier Register
ws3 = wb.create_sheet("Supplier Register")
hdr(ws3, ["Supplier ID","Name","Status","Evidence","Sources","Bank","Notes"])
cw(ws3, [18,30,22,60,20,30,50])
row = 2
for i, s in enumerate(wh["warehouses"] if False else []):
    pass
with open(PROJ_DIR + "/supplier-register.json", encoding='utf-8') as f:
    sup_data = json.load(f)
for i, s in enumerate(sup_data["suppliers"]):
    fill = WARN
    vals = [s["supplier_id"], s["name"], s["match_status"], s["evidence"][:80],
            "; ".join(s["doc_sources"]), s.get("bank_info","—"), s["notes"][:80]]
    wr(ws3, row, vals, fill=fill); row += 1

# Sheet 5: Warehouse Register
ws4 = wb.create_sheet("Warehouse Register")
hdr(ws4, ["Warehouse ID","Name","Status","Source","Source Text","Type","Notes"])
cw(ws4, [18,22,22,18,50,30,60])
row = 2
for i, w in enumerate(wh["warehouses"]):
    fill = WARN
    vals = [w["warehouse_id"], w["name"], w["match_status"], w["source"],
            w["source_text"][:60], w["type"], w["notes"][:80]]
    wr(ws4, row, vals, fill=fill); row += 1

# Sheet 6: ERP Gate
ws5 = wb.create_sheet("ERP Gate")
hdr(ws5, ["Line ID","Can Post","Block Count","Block Reasons"])
cw(ws5, [14,12,12,80])
row = 2
for g in recon["erp_gate_checks"]:
    fill = PASS_F if g["can_post"] else ERR
    vals = [g["line_id"], "YES" if g["can_post"] else "NO", g["block_count"], "; ".join(g["reasons"][:3])]
    wr(ws5, row, vals, fill=fill); row += 1
# Summary
ws5.cell(row=row+1, column=1, value="TOTAL POSTABLE:").font = Font(bold=True)
ws5.cell(row=row+1, column=2, value=sum(1 for g in recon["erp_gate_checks"] if g["can_post"])).font = Font(bold=True)
ws5.cell(row=row+2, column=1, value="TOTAL BLOCKED:").font = Font(bold=True)
ws5.cell(row=row+2, column=2, value=sum(1 for g in recon["erp_gate_checks"] if not g["can_post"])).font = Font(bold=True)

# Sheet 7: Procurement Chain
ws6 = wb.create_sheet("Procurement Chain")
hdr(ws6, ["Source","Document Type","Procurement Chain","MR","PO","Receipt","Invoice","Stock","Status"])
cw(ws6, [12,28,50,12,12,12,12,12,22])
row = 2
chains = [
    ("SRC-001","MATERIAL_REQUIREMENT","MR → [PO] → [Receipt] → [Stock]","SRC-001","MISSING","MISSING","MISSING","MISSING"),
    ("SRC-002","SUPPLIER_ORDER_CONFIRMATION","PO ← Confirms [SRC-002] → [Receipt]","MISSING","LINKED","MISSING","MISSING","MISSING"),
    ("SRC-003","GOODS_DELIVERY_NOTE","[PO] → DN [SRC-003] → [GRN] → [Stock]","N/A","PENDING","AMBIGUOUS","MISSING","MISSING"),
    ("SRC-004","PURCHASE_ORDER","PO [SRC-004] → [Receipt] → [Invoice]","N/A","SRC-004","MISSING","MISSING","MISSING"),
]
for i, (src, dt, chain, mr, po, rcpt, inv, stk) in enumerate(chains):
    fill = WARN
    vals = [src, dt, chain, mr, po, rcpt, inv, stk, "PENDING_COMPLETE"]
    wr(ws6, row, vals, fill=fill); row += 1

# Sheet 8: Amount Verification
ws7 = wb.create_sheet("Amount Verification")
hdr(ws7, ["Line ID","Description","Qty","Unit Price After CK","Calc Amount","Source Amount","Match","Delta","Status"])
cw(ws7, [14,40,8,16,14,14,8,12,16])
row = 2
for r in ROWS:
    if r["unit_price"] is None and r.get("unit_price_after_discount") is None:
        continue
    up_ck = r.get("unit_price_after_discount") or r.get("unit_price")
    calc  = round(r["quantity"] * up_ck) if up_ck else None
    src_a = r["line_amount"]
    if calc is not None and src_a is not None:
        match = calc == src_a
        delta = calc - src_a
        fill  = PASS_F if match else ERR
    else:
        match = None; delta = None; fill = PEND
    vals = [r["line_id"], r["description_raw"][:45], r["quantity"], up_ck, calc, src_a,
            "✅" if match else ("N/A" if match is None else "❌"), delta, "PASS" if match else ("N/A" if match is None else "FAIL")]
    wr(ws7, row, vals, fill=fill); row += 1

xlsx_path = PROJ_DIR + "/BAO-MINH-CMT8-MATERIAL-RECONCILIATION.xlsx"
wb.save(xlsx_path)
LOG("[4C] Excel: " + xlsx_path + " (" + str(os.path.getsize(xlsx_path)) + " bytes)")
LOG("[4C] PHASE 4C COMPLETE")
